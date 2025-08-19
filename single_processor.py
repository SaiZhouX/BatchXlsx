import pandas as pd
import os
from pathlib import Path
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class SingleProcessor:
    """单个Excel文件处理器"""
    
    def __init__(self, output_folder="output"):
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(exist_ok=True)
        self.logger = logging.getLogger(__name__)
    
    def process_single_file(self, file_path):
        """处理单个Excel文件"""
        try:
            self.logger.info(f"开始处理文件: {os.path.basename(file_path)}")
            
            # 读取Excel文件
            df = pd.read_excel(file_path)
            self.logger.info(f"成功读取文件，共 {len(df)} 行数据")
            
            if df.empty:
                self.logger.warning("文件为空，无法处理")
                return None
            
            # 数据清理
            df_cleaned = self.clean_data(df)
            
            # 生成报告
            report_path = self.generate_detailed_report(df_cleaned, file_path)
            
            self.logger.info(f"单个文件分析完成: {os.path.basename(file_path)}")
            return report_path
            
        except Exception as e:
            self.logger.error(f"处理文件 {file_path} 时出错: {str(e)}")
            return None
    
    def clean_data(self, df):
        """清理数据"""
        try:
            # 记录原始数据信息
            original_rows = len(df)
            original_cols = len(df.columns)
            
            # 删除完全为空的行
            df_cleaned = df.dropna(how='all').copy()
            
            # 删除无用的列（如Unnamed列）
            unnamed_cols = [col for col in df_cleaned.columns if 'Unnamed' in str(col)]
            if unnamed_cols:
                df_cleaned = df_cleaned.drop(columns=unnamed_cols)
                self.logger.info(f"删除了无用列: {unnamed_cols}")
            
            # 删除完全为空的列
            df_cleaned = df_cleaned.dropna(axis=1, how='all')
            
            cleaned_rows = len(df_cleaned)
            cleaned_cols = len(df_cleaned.columns)
            
            self.logger.info(f"数据清理完成: 原有 {original_rows} 行 {original_cols} 列，清理后 {cleaned_rows} 行 {cleaned_cols} 列")
            
            return df_cleaned
            
        except Exception as e:
            self.logger.error(f"数据清理失败: {str(e)}")
            return df
    
    def generate_detailed_report(self, df, source_file_path):
        """生成详细分析报告"""
        try:
            # 生成报告文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            source_filename = Path(source_file_path).stem
            report_filename = f"详细分析报告_{source_filename}_{timestamp}.xlsx"
            report_path = self.output_folder / report_filename
            
            # 创建Excel写入器
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                # 生成数据统计（放在第一个工作表）
                stats_data = self.generate_statistics(df, source_file_path)
                stats_df = pd.DataFrame(list(stats_data.items()), columns=['统计项', '值'])
                stats_df.to_excel(writer, sheet_name='分析摘要', index=False)
                
                # 写入清理后的数据
                df.to_excel(writer, sheet_name='详细数据', index=False)
                
                # 如果数据量不大，添加数据预览
                if len(df) <= 1000:
                    preview_df = df.head(100)  # 显示前100行
                    preview_df.to_excel(writer, sheet_name='数据预览', index=False)
                
                # 格式化工作表
                self.format_worksheet(writer.sheets['分析摘要'], stats_df)
                self.format_worksheet(writer.sheets['详细数据'], df)
                if len(df) <= 1000:
                    self.format_worksheet(writer.sheets['数据预览'], preview_df)
            
            self.logger.info(f"已生成详细分析报告: {report_filename}")
            return report_path
            
        except Exception as e:
            self.logger.error(f"生成详细分析报告失败: {str(e)}")
            return None
    
    def generate_statistics(self, df, source_file_path):
        """生成数据统计信息"""
        stats = {}
        
        try:
            # 添加报告标题
            stats['=== Excel文件分析报告 ==='] = '='*30
            stats[''] = ''  # 空行
            
            # 基本统计
            stats['源文件名称'] = os.path.basename(source_file_path)
            stats['分析时间'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            stats['数据总行数'] = len(df)
            stats['数据总列数'] = len(df.columns)
            stats['有效数据行数'] = len(df.dropna(how='all'))
            
            # 列统计
            stats['数值列数量'] = len(df.select_dtypes(include=['number']).columns)
            stats['文本列数量'] = len(df.select_dtypes(include=['object']).columns)
            stats['日期列数量'] = len(df.select_dtypes(include=['datetime']).columns)
            
            # 数据质量统计
            total_cells = len(df) * len(df.columns)
            empty_cells = df.isnull().sum().sum()
            stats['空值单元格数'] = empty_cells
            if total_cells > 0:
                stats['数据完整率'] = f"{((total_cells - empty_cells) / total_cells * 100):.1f}%"
            else:
                stats['数据完整率'] = "0.0%"
            
            # 如果有特定的业务列，进行业务统计
            self.add_business_statistics(df, stats)
            
        except Exception as e:
            self.logger.error(f"生成统计信息失败: {str(e)}")
            stats['错误信息'] = str(e)
        
        return stats
    
    def add_business_statistics(self, df, stats):
        """添加业务相关统计"""
        try:
            # 检查是否有Bug相关的列
            bug_columns = [col for col in df.columns if any(keyword in col.lower() for keyword in ['bug', '缺陷', '问题', '错误'])]
            if bug_columns:
                stats['Bug相关列数'] = len(bug_columns)
            
            # 检查是否有级别相关的列
            level_columns = [col for col in df.columns if any(keyword in col.lower() for keyword in ['级别', 'level', '等级', 'priority', '严重', 'severity'])]
            if level_columns:
                level_col = level_columns[0]
                level_counts = df[level_col].value_counts()
                for level, count in level_counts.items():
                    if pd.notna(level):
                        stats[f'{level}数量'] = count
            
            # 检查是否有状态相关的列
            status_columns = [col for col in df.columns if any(keyword in col.lower() for keyword in ['状态', 'status', '修复', 'fixed', '解决', 'resolved'])]
            if status_columns:
                status_col = status_columns[0]
                status_counts = df[status_col].value_counts()
                for status, count in status_counts.items():
                    if pd.notna(status):
                        stats[f'{status}数量'] = count
            
            # 检查是否有类型相关的列
            type_columns = [col for col in df.columns if any(keyword in col.lower() for keyword in ['类型', 'type', '分类', 'category'])]
            if type_columns:
                type_col = type_columns[0]
                type_counts = df[type_col].value_counts()
                for bug_type, count in type_counts.items():
                    if pd.notna(bug_type):
                        stats[f'{bug_type}数量'] = count
                        
        except Exception as e:
            self.logger.error(f"添加业务统计失败: {str(e)}")
    
    def format_worksheet(self, worksheet, df):
        """格式化工作表"""
        try:
            # 设置标题行样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 应用标题行样式
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 设置列宽
            for col_num, column in enumerate(df.columns, 1):
                column_letter = worksheet.cell(row=1, column=col_num).column_letter
                max_length = max(
                    len(str(column)),
                    max([len(str(df.iloc[row, col_num-1])) for row in range(min(len(df), 100))], default=0)
                )
                adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 为数据行添加边框
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).border = thin_border
            
        except Exception as e:
            self.logger.error(f"格式化工作表失败: {str(e)}")
    
    def process_multiple_files(self, file_paths):
        """处理多个单独的Excel文件"""
        results = []
        
        for file_path in file_paths:
            result = self.process_single_file(file_path)
            if result:
                results.append(result)
        
        return results